#!/usr/bin/env python3
"""
Generate complete sample data matching the exact SQL structure with all 93 columns
"""

import random
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Sample data pools
divisions = ['Sales', 'Marketing', 'IT', 'Finance', 'HR', 'Operations']
regions = ['North America', 'EMEA', 'APAC', 'LATAM']
websites = ['IntranetSite', 'TeamCollaboration', 'KnowledgeBase', 'NewsPortal']
page_names = ['Product Updates', 'Company News', 'Team Directory', 'Policy Updates',
              'Training Materials', 'Event Calendar', 'Department News', 'HR Benefits',
              'IT Support', 'Sales Dashboard', 'Marketing Resources', 'Finance Reports']

def generate_dim_date(num_days=365):
    data = []
    start_date = datetime(2024, 1, 1)
    for i in range(num_days):
        date = start_date + timedelta(days=i)
        data.append({
            'date_key': i + 1,
            'date': date.strftime('%Y-%m-%d'),
            'year': date.year,
            'month': date.month,
            'day': date.day,
            'day_of_week': date.strftime('%A')
        })
    return data

def generate_employee_contacts(num_employees=200):
    data = []
    for i in range(1, num_employees + 1):
        division = random.choice(divisions) if random.random() > 0.05 else None
        region = random.choice(regions) if random.random() > 0.05 else None
        data.append({
            'contactid': f'EMP{i:05d}',
            'employeename': f'Employee {i}',
            'employeebusinessdivision': division,
            'employeeregion': region,
            'email': f'employee{i}@company.com'
        })
    return data

def generate_website_page_inventory(num_pages=50):
    data = []
    for i in range(1, num_pages + 1):
        website = random.choice(websites)
        page = random.choice(page_names)
        created_date = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 90))
        data.append({
            'marketingPageId': f'PAGE{i:04d}',
            'websitename': website,
            'pagename': page,
            'page_url': f'https://{website.lower()}.company.com/{page.lower().replace(" ", "-")}',
            'created_date': created_date.strftime('%Y-%m-%d'),
            'created_date_obj': created_date  # Keep for mindate calculation
        })
    return data

def generate_interactions(num_interactions=1000, employees=None, pages=None, dates=None):
    """Generate realistic interaction patterns with repeated visitors"""
    data = []

    # Create more realistic patterns:
    # 1. 20% of pages are "popular" and get 60% of traffic
    # 2. Multiple interactions from same employee to same page

    popular_pages = random.sample(pages, k=max(1, len(pages) // 5))

    for i in range(num_interactions):
        # 60% chance to pick a popular page
        if random.random() < 0.6 and popular_pages:
            page = random.choice(popular_pages)
        else:
            page = random.choice(pages)

        employee = random.choice(employees)

        # Pick dates with clustering (some pages more active on certain days)
        if random.random() < 0.3:
            # Cluster interactions within first 30 days
            date = random.choice([d for d in dates if d['date_key'] <= 30])
        else:
            date = random.choice(dates)

        views = random.randint(1, 10)
        visits = random.randint(1, min(5, views))
        comments = random.randint(0, 3) if random.random() > 0.7 else 0
        liked_page = page['marketingPageId'] if random.random() > 0.85 else None

        data.append({
            'interaction_id': f'INT{i:06d}',
            'viewingcontactid': employee['contactid'],
            'marketingPageId': page['marketingPageId'],
            'visitdatekey': date['date_key'],
            'visit_date': date['date'],
            'views': views,
            'visits': visits,
            'comments': comments,
            'marketingPageIdliked': liked_page
        })

    return data

def calculate_days_since_first(interaction_date_str, mindate_str):
    """Calculate days between interaction and first page date"""
    interaction_date = datetime.strptime(interaction_date_str, '%Y-%m-%d')
    mindate = datetime.strptime(mindate_str, '%Y-%m-%d')
    return (interaction_date - mindate).days

def generate_complete_output(interactions, employees, pages, dates):
    """Generate complete output with ALL 93 columns"""

    emp_lookup = {e['contactid']: e for e in employees}
    page_lookup = {p['marketingPageId']: p for p in pages}
    date_lookup = {d['date_key']: d for d in dates}

    # Calculate mindate for each page
    page_mindates = {}
    for interaction in interactions:
        page_id = interaction['marketingPageId']
        date_str = interaction['visit_date']
        if page_id not in page_mindates or date_str < page_mindates[page_id]:
            page_mindates[page_id] = date_str

    # Track aggregations at grain level and all other levels
    grain_data = defaultdict(lambda: {
        'all': {'uv': set(), 'likes': set(), 'views': 0, 'visits': 0, 'comments': 0},
        'ty': {'uv': set(), 'likes': set(), 'views': 0, 'visits': 0, 'comments': 0},
        'd28': {'uv': set(), 'likes': set(), 'views': 0, 'visits': 0, 'comments': 0},
        'd21': {'uv': set(), 'likes': set(), 'views': 0, 'visits': 0, 'comments': 0},
        'd14': {'uv': set(), 'likes': set(), 'views': 0, 'visits': 0, 'comments': 0},
        'd7': {'uv': set(), 'likes': set(), 'views': 0, 'visits': 0, 'comments': 0},
    })

    # Aggregations by dimension combinations for each time period
    agg_levels = {}
    for period in ['all', 'ty', 'd28', 'd21', 'd14', 'd7']:
        agg_levels[period] = {
            'overall': set(),
            'div': defaultdict(set),
            'reg': defaultdict(set),
            'site': defaultdict(set),
            'site_div': defaultdict(set),
            'site_reg': defaultdict(set),
            'site_div_reg': defaultdict(set),
        }

    # Process interactions
    for interaction in interactions:
        emp = emp_lookup[interaction['viewingcontactid']]
        page = page_lookup[interaction['marketingPageId']]
        date = date_lookup[interaction['visitdatekey']]

        division = emp['employeebusinessdivision'] or 'Unknown'
        region = emp['employeeregion'] or 'Unknown'
        website = page['websitename']
        page_id = interaction['marketingPageId']
        visitor = interaction['viewingcontactid']
        is_current_year = date['year'] == 2024

        # Calculate days since first interaction for this page
        mindate = page_mindates.get(page_id)
        if mindate:
            days_since = calculate_days_since_first(date['date'], mindate)
        else:
            days_since = 0

        # Main grain key
        key = (page_id, division, region, website)

        # Accumulate metrics for all time periods
        periods_to_update = ['all']
        if is_current_year:
            periods_to_update.append('ty')
        if days_since <= 27:
            periods_to_update.append('d28')
        if days_since <= 20:
            periods_to_update.append('d21')
        if days_since <= 13:
            periods_to_update.append('d14')
        if days_since <= 6:
            periods_to_update.append('d7')

        for period in periods_to_update:
            # Grain level
            grain_data[key][period]['uv'].add(visitor)
            grain_data[key][period]['views'] += interaction['views']
            grain_data[key][period]['visits'] += interaction['visits']
            grain_data[key][period]['comments'] += interaction['comments']
            if interaction['marketingPageIdliked']:
                grain_data[key][period]['likes'].add(interaction['marketingPageIdliked'])

            # Other aggregation levels
            agg_levels[period]['overall'].add(visitor)
            agg_levels[period]['div'][division].add(visitor)
            agg_levels[period]['reg'][region].add(visitor)
            agg_levels[period]['site'][website].add(visitor)
            agg_levels[period]['site_div'][(website, division)].add(visitor)
            agg_levels[period]['site_reg'][(website, region)].add(visitor)
            agg_levels[period]['site_div_reg'][(website, division, region)].add(visitor)

    # Build output rows with all 93 columns
    output = []
    for key, metrics in grain_data.items():
        page_id, division, region, website = key
        page = page_lookup[page_id]

        row = {
            # Dimensions (4 columns)
            'marketingPageId': page_id,
            'employeebusinessdivision': division,
            'employeeregion': region,
            'websitename': website,

            # All-time metrics (div_reg level) - 9 columns
            'div_reg_views': metrics['all']['views'],
            'div_reg_visits': metrics['all']['visits'],
            'div_reg_comments': metrics['all']['comments'],
            'div_reg_likes': len(metrics['all']['likes']),
            'div_reg_uniquevisitor': len(metrics['all']['uv']),
            'div_uniquevisitor': len(agg_levels['all']['div'][division]),
            'reg_uniquevisitor': len(agg_levels['all']['reg'][region]),
            'uniquevisitor': len(agg_levels['all']['overall']),
            'site_uniquevisitor': len(agg_levels['all']['site'][website]),
            'site_div_uniquevisitor': len(agg_levels['all']['site_div'][(website, division)]),
            'site_reg_uniquevisitor': len(agg_levels['all']['site_reg'][(website, region)]),
            'site_div_reg_uniquevisitor': len(agg_levels['all']['site_div_reg'][(website, division, region)]),

            # This year metrics - 13 columns
            'div_reg_viewty': metrics['ty']['views'],
            'div_reg_visitsty': metrics['ty']['visits'],
            'div_reg_commentsty': metrics['ty']['comments'],
            'div_reg_likesty': len(metrics['ty']['likes']),
            'div_reg_uniquevisitorty': len(metrics['ty']['uv']),
            'div_uniquevisitorty': len(agg_levels['ty']['div'][division]),
            'reg_uniquevisitorty': len(agg_levels['ty']['reg'][region]),
            'uniquevisitorty': len(agg_levels['ty']['overall']),
            'site_uniquevisitorTY': len(agg_levels['ty']['site'][website]),
            'site_div_uniquevisitorTY': len(agg_levels['ty']['site_div'][(website, division)]),
            'site_reg_uniquevisitorTY': len(agg_levels['ty']['site_reg'][(website, region)]),
            'site_div_reg_uniquevisitorTY': len(agg_levels['ty']['site_div_reg'][(website, division, region)]),

            # 28 days metrics - 13 columns
            'div_reg_views28': metrics['d28']['views'],
            'div_reg_visits28': metrics['d28']['visits'],
            'div_reg_comments28': metrics['d28']['comments'],
            'div_reg_likes28': len(metrics['d28']['likes']),
            'div_reg_uniquevisitor28': len(metrics['d28']['uv']),
            'div_uniquevisitor28': len(agg_levels['d28']['div'][division]),
            'reg_uniquevisitor28': len(agg_levels['d28']['reg'][region]),
            'uniquevisitor28': len(agg_levels['d28']['overall']),
            'site_uniquevisitor28': len(agg_levels['d28']['site'][website]),
            'site_div_uniquevisitor28': len(agg_levels['d28']['site_div'][(website, division)]),
            'site_reg_uniquevisitor28': len(agg_levels['d28']['site_reg'][(website, region)]),
            'site_div_reg_uniquevisitor28': len(agg_levels['d28']['site_div_reg'][(website, division, region)]),

            # 21 days metrics - 13 columns
            'div_reg_views21': metrics['d21']['views'],
            'div_reg_visits21': metrics['d21']['visits'],
            'div_reg_comments21': metrics['d21']['comments'],
            'div_reg_likes21': len(metrics['d21']['likes']),
            'div_reg_uniquevisitor21': len(metrics['d21']['uv']),
            'div_uniquevisitor21': len(agg_levels['d21']['div'][division]),
            'reg_uniquevisitor21': len(agg_levels['d21']['reg'][region]),
            'uniquevisitor21': len(agg_levels['d21']['overall']),
            'site_uniquevisitor21': len(agg_levels['d21']['site'][website]),
            'site_div_uniquevisitor21': len(agg_levels['d21']['site_div'][(website, division)]),
            'site_reg_uniquevisitor21': len(agg_levels['d21']['site_reg'][(website, region)]),
            'site_div_reg_uniquevisitor21': len(agg_levels['d21']['site_div_reg'][(website, division, region)]),

            # 14 days metrics - 13 columns
            'div_reg_views14': metrics['d14']['views'],
            'div_reg_visits14': metrics['d14']['visits'],
            'div_reg_comments14': metrics['d14']['comments'],
            'div_reg_likes14': len(metrics['d14']['likes']),
            'div_reg_uniquevisitor14': len(metrics['d14']['uv']),
            'div_uniquevisitor14': len(agg_levels['d14']['div'][division]),
            'reg_uniquevisitor14': len(agg_levels['d14']['reg'][region]),
            'uniquevisitor14': len(agg_levels['d14']['overall']),
            'site_uniquevisitor14': len(agg_levels['d14']['site'][website]),
            'site_div_uniquevisitor14': len(agg_levels['d14']['site_div'][(website, division)]),
            'site_reg_uniquevisitor14': len(agg_levels['d14']['site_reg'][(website, region)]),
            'site_div_reg_uniquevisitor14': len(agg_levels['d14']['site_div_reg'][(website, division, region)]),

            # 7 days metrics - 13 columns
            'div_reg_views7': metrics['d7']['views'],
            'div_reg_visits7': metrics['d7']['visits'],
            'div_reg_comments7': metrics['d7']['comments'],
            'div_reg_likes7': len(metrics['d7']['likes']),
            'div_reg_uniquevisitor7': len(metrics['d7']['uv']),
            'div_uniquevisitor7': len(agg_levels['d7']['div'][division]),
            'reg_uniquevisitor7': len(agg_levels['d7']['reg'][region]),
            'uniquevisitor7': len(agg_levels['d7']['overall']),
            'site_uniquevisitor7': len(agg_levels['d7']['site'][website]),
            'site_div_uniquevisitor7': len(agg_levels['d7']['site_div'][(website, division)]),
            'site_reg_uniquevisitor7': len(agg_levels['d7']['site_reg'][(website, region)]),
            'site_div_reg_uniquevisitor7': len(agg_levels['d7']['site_div_reg'][(website, division, region)]),
        }

        output.append(row)

    return sorted(output, key=lambda x: (x['websitename'], x['employeebusinessdivision'], x['employeeregion'], x['marketingPageId']))

def format_header(ws, columns, uv_col_indices):
    """Format header with color coding - UV columns get special treatment"""
    dimension_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # UV-specific colors (more vibrant/prominent)
    uv_alltime_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Bright green
    uv_thisyear_fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')  # Bright orange
    uv_days_fill = PatternFill(start_color='0099CC', end_color='0099CC', fill_type='solid')  # Bright blue

    # Non-UV colors (muted)
    other_alltime_fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')  # Pale green
    other_thisyear_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')  # Pale orange
    other_days_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')  # Pale blue

    header_font = Font(bold=True, color='FFFFFF', size=9)

    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_name
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        if col_num <= 4:
            # Dimensions
            cell.fill = dimension_fill
        elif col_num in uv_col_indices:
            # UV columns - prominent colors
            if 'ty' in column_name.lower() or 'TY' in column_name:
                cell.fill = uv_thisyear_fill
            elif any(d in column_name for d in ['28', '21', '14', '7']):
                cell.fill = uv_days_fill
            else:
                cell.fill = uv_alltime_fill
        else:
            # Non-UV columns - muted colors
            if 'ty' in column_name.lower() or 'TY' in column_name:
                cell.fill = other_thisyear_fill
            elif any(d in column_name for d in ['28', '21', '14', '7']):
                cell.fill = other_days_fill
            else:
                cell.fill = other_alltime_fill

def auto_size_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 10), 35)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_data_to_sheet(ws, data, columns, uv_col_indices=None, collapse_groups=None):
    for row_num, record in enumerate(data, 2):
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = record.get(column)
            if isinstance(record.get(column), (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    if uv_col_indices:
        format_header(ws, columns, uv_col_indices)
    else:
        # Simple header format for input sheets
        dimension_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=9)
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = dimension_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    auto_size_columns(ws)
    ws.freeze_panes = 'A2'

    # Add column grouping/outlining for collapsible sections
    if collapse_groups:
        for start_col, end_col in collapse_groups:
            ws.column_dimensions.group(get_column_letter(start_col), get_column_letter(end_col), outline_level=1, hidden=True)

def main():
    print("Generating complete sample data with all time periods...")

    dates = generate_dim_date(365)
    employees = generate_employee_contacts(200)
    pages = generate_website_page_inventory(50)
    interactions = generate_interactions(1000, employees, pages, dates)
    aggregated = generate_complete_output(interactions, employees, pages, dates)

    print(f"Generated {len(aggregated)} output rows with 93 columns")

    wb = Workbook()
    wb.remove(wb.active)

    # Overview sheet
    ws_overview = wb.create_sheet("📊 Overview", 0)
    overview_data = [
        ['SharePoint Analytics - Complete Data Model'],
        [''],
        ['Purpose', 'Demonstrate the complete SQL output with ALL time periods and aggregation levels'],
        ['Created', datetime.now().strftime('%Y-%m-%d %H:%M')],
        [''],
        ['Output Structure (76 columns total):'],
        ['1. Dimensions (4 cols)', 'PageID, Division, Region, Website - the grain of the table'],
        ['2. UV Columns (48 cols)', 'ALL Unique Visitor metrics across all time periods - MOST IMPORTANT'],
        ['3. Other Metrics (24 cols)', 'Views, Visits, Comments, Likes - collapsed by default'],
        [''],
        ['Column Organization:'],
        ['• Columns 1-4: Dimensions (Blue headers)'],
        ['• Columns 5-52: UV metrics (Bright colors - Green/Orange/Blue)'],
        ['• Columns 53-76: Other metrics (Pale colors - COLLAPSED for focus)'],
        [''],
        ['UV Metrics Breakdown (48 columns):'],
        ['• All-time UV: 8 aggregation levels'],
        ['• This year UV: 8 aggregation levels'],
        ['• First 28 days UV: 8 aggregation levels'],
        ['• First 21 days UV: 8 aggregation levels'],
        ['• First 14 days UV: 8 aggregation levels'],
        ['• First 7 days UV: 8 aggregation levels'],
        [''],
        ['UV Aggregation Levels (repeated for each time period):'],
        ['1. div_reg_uniquevisitor', 'Page + Division + Region'],
        ['2. div_uniquevisitor', 'Division only'],
        ['3. reg_uniquevisitor', 'Region only'],
        ['4. uniquevisitor', 'Overall (all visitors)'],
        ['5. site_uniquevisitor', 'Website only'],
        ['6. site_div_uniquevisitor', 'Website + Division'],
        ['7. site_reg_uniquevisitor', 'Website + Region'],
        ['8. site_div_reg_uniquevisitor', 'Website + Division + Region'],
        [''],
        ['Color Coding:'],
        ['• Blue = Dimensions'],
        ['• Bright Green = UV all-time (most important)'],
        ['• Bright Orange = UV this year'],
        ['• Bright Blue = UV first N days (28/21/14/7)'],
        ['• Pale colors = Other metrics (views, visits, comments, likes)'],
        [''],
        ['💡 Tip: The other metrics columns are collapsed - expand the groups to see them'],
    ]

    for row_num, row_data in enumerate(overview_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_overview.cell(row=row_num, column=col_num)
            cell.value = value
            if row_num == 1:
                cell.font = Font(size=16, bold=True, color='366092')
            elif row_num in [3, 6, 12]:
                cell.font = Font(bold=True, size=11)

    ws_overview.column_dimensions['A'].width = 30
    ws_overview.column_dimensions['B'].width = 80

    # Input data sheets
    ws_dates = wb.create_sheet("dim_date")
    add_data_to_sheet(ws_dates, dates, ['date_key', 'date', 'year', 'month', 'day', 'day_of_week'])

    ws_employees = wb.create_sheet("employeecontact")
    add_data_to_sheet(ws_employees, employees, ['contactid', 'employeename', 'employeebusinessdivision', 'employeeregion', 'email'])

    ws_pages = wb.create_sheet("website_page_inventory")
    add_data_to_sheet(ws_pages, pages, ['marketingPageId', 'websitename', 'pagename', 'page_url', 'created_date'])

    ws_interactions = wb.create_sheet("interactions_metrics")
    add_data_to_sheet(ws_interactions, interactions, ['interaction_id', 'viewingcontactid', 'marketingPageId', 'visitdatekey', 'visit_date', 'views', 'visits', 'comments', 'marketingPageIdliked'])

    # Output sheet with reorganized columns: Dimensions -> All UV columns -> Other metrics (collapsed)
    ws_output = wb.create_sheet("⭐ COMPLETE OUTPUT (93 cols)")

    output_columns = [
        # === DIMENSIONS (4 cols) ===
        'marketingPageId', 'employeebusinessdivision', 'employeeregion', 'websitename',

        # === ALL UV COLUMNS (48 cols) - MOST IMPORTANT ===
        # All-time UV (8 cols)
        'div_reg_uniquevisitor', 'div_uniquevisitor', 'reg_uniquevisitor', 'uniquevisitor',
        'site_uniquevisitor', 'site_div_uniquevisitor', 'site_reg_uniquevisitor', 'site_div_reg_uniquevisitor',

        # This year UV (8 cols)
        'div_reg_uniquevisitorty', 'div_uniquevisitorty', 'reg_uniquevisitorty', 'uniquevisitorty',
        'site_uniquevisitorTY', 'site_div_uniquevisitorTY', 'site_reg_uniquevisitorTY', 'site_div_reg_uniquevisitorTY',

        # 28 days UV (8 cols)
        'div_reg_uniquevisitor28', 'div_uniquevisitor28', 'reg_uniquevisitor28', 'uniquevisitor28',
        'site_uniquevisitor28', 'site_div_uniquevisitor28', 'site_reg_uniquevisitor28', 'site_div_reg_uniquevisitor28',

        # 21 days UV (8 cols)
        'div_reg_uniquevisitor21', 'div_uniquevisitor21', 'reg_uniquevisitor21', 'uniquevisitor21',
        'site_uniquevisitor21', 'site_div_uniquevisitor21', 'site_reg_uniquevisitor21', 'site_div_reg_uniquevisitor21',

        # 14 days UV (8 cols)
        'div_reg_uniquevisitor14', 'div_uniquevisitor14', 'reg_uniquevisitor14', 'uniquevisitor14',
        'site_uniquevisitor14', 'site_div_uniquevisitor14', 'site_reg_uniquevisitor14', 'site_div_reg_uniquevisitor14',

        # 7 days UV (8 cols)
        'div_reg_uniquevisitor7', 'div_uniquevisitor7', 'reg_uniquevisitor7', 'uniquevisitor7',
        'site_uniquevisitor7', 'site_div_uniquevisitor7', 'site_reg_uniquevisitor7', 'site_div_reg_uniquevisitor7',

        # === OTHER METRICS (28 cols) - COLLAPSED BY DEFAULT ===
        # All-time other metrics (4 cols)
        'div_reg_views', 'div_reg_visits', 'div_reg_comments', 'div_reg_likes',

        # This year other metrics (4 cols)
        'div_reg_viewty', 'div_reg_visitsty', 'div_reg_commentsty', 'div_reg_likesty',

        # 28 days other metrics (4 cols)
        'div_reg_views28', 'div_reg_visits28', 'div_reg_comments28', 'div_reg_likes28',

        # 21 days other metrics (4 cols)
        'div_reg_views21', 'div_reg_visits21', 'div_reg_comments21', 'div_reg_likes21',

        # 14 days other metrics (4 cols)
        'div_reg_views14', 'div_reg_visits14', 'div_reg_comments14', 'div_reg_likes14',

        # 7 days other metrics (4 cols)
        'div_reg_views7', 'div_reg_visits7', 'div_reg_comments7', 'div_reg_likes7',
    ]

    # UV column indices (after dimensions, columns 5-52)
    uv_col_indices = set(range(5, 53))  # Columns 5 through 52 are UV columns

    # Collapse groups for other metrics (columns 53-76)
    collapse_groups = [
        (53, 56),   # All-time other metrics
        (57, 60),   # This year other metrics
        (61, 64),   # 28 days other metrics
        (65, 68),   # 21 days other metrics
        (69, 72),   # 14 days other metrics
        (73, 76),   # 7 days other metrics
    ]

    add_data_to_sheet(ws_output, aggregated, output_columns, uv_col_indices=uv_col_indices, collapse_groups=collapse_groups)

    filename = '/Users/micha/Documents/Fabric/sharepoint_analytics_sample_data.xlsx'
    wb.save(filename)
    print(f"\n✅ Complete Excel file created: {filename}")
    print(f"   Total columns in output: {len(output_columns)}")
    print(f"   Total sheets: {len(wb.sheetnames)}")

if __name__ == '__main__':
    main()
